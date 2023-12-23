import openpyxl

def parse(file) -> list:
    data = []
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        row = [ws.cell(row = i, column = x).value for x in range(1,ws.max_column + 1)]
        row = list(map(str, row))
        name = " ".join([row[0],row[1],row[2]])
        departure_city = row[3]
        destination_city = row[4]
        departure_date = row[5].split(" ")[0]
        back_date = row[6] = row[6].split(" ")[0]
        
        business_trip = {
            "name": name,
            "departure_city": departure_city,
            "destination_city": destination_city,
            "departure_date": departure_date,
            "back_date": back_date
        }
        data.append(business_trip)
    return data
    
    